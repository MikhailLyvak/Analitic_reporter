from excel_writer_class import Report
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Border, Side, NamedStyle, Font, Alignment
from openpyxl import load_workbook


def report_4(report1_data: list, report2_data: list, report3_data: list, month: int) -> None:
    titles1 = [
            "Менеджер",
            "Дата",
            "Мінімальний продаж"
        ]
    titles2 = [
            "Менеджер",
            "Дата",
            "Максимальний продаж"
        ]
    titles3 = [
            "Менеджер",
            "Чистий прибуток"
        ]
    t2 = report2_data
    t3 = report3_data
    print(t2, t3)

    # Создаем первый отчет
    res1 = Report(data=report1_data, report_name=f"Мінамальні продажі по менеджерам за {month.lower()} місяць")
    res1.write_data(titles=titles1)
    res1.number_columns()
    res1.zagolovok(place=["B2", "E2"], height=50)
    res1.design_titles()
    res1.borders()
    res1.ws.column_dimensions['C'].width = 30
    res1.ws.column_dimensions['D'].width = 17
    res1.ws.column_dimensions['E'].width = 17
    res1.align_centre()
    res1.name_align_left(4, 3, 3)
    res1.money_format(4, 4)
    
    # Сохраняем первый отчет в файл
    res1.save("Report4.xlsx")
    
    # Загружаем существующую книгу
    res2 = Report(data=report2_data, report_name=f"Максимальні продажі по менеджерам за {month.lower()} місяць")
    res2.load_workbook("Report4.xlsx")
    
    # Находим первую пустую строку
    row_start = len(res2.ws['A']) + 1
    
    # Записываем данные второго отчета
    row_start += 3  # Добавляем строку для разделения отчетов
    res2.write_data(titles=titles2, row_start=row_start)
    res2.number_columns(row_start=row_start)
    res2.zagolovok(place=["B14", "E14"], height=50)
    res2.design_titles(row_number=row_start)
    res2.borders(x_start=row_start)
    res2.align_centre()
    res2.name_align_left(4, 3, 3)
    res2.money_format(4, 4)
    
    # Сохраняем книгу с обоими отчетами
    res2.save("Report4.xlsx")

    res3 = Report(data=report3_data, report_name=f"Чистий прибуток по менеджерам за {month.lower()} місяць")
    res3.load_workbook("Report4.xlsx")
    
    # Находим первую пустую строку
    row_start = len(res3.ws['A']) + 1
    
    # Записываем данные второго отчета
    row_start += 3
    res3.write_data(titles=titles3, row_start=row_start)
    res3.number_columns(row_start=row_start)
    res3.zagolovok(place=["B26", "D26"], height=50)
    res3.align_centre()
    res3.borders(x_start=row_start)
    res3.name_align_left(4, 3, 3)
    res3.money_format(4, 4)
    res3.design_titles(row_number=row_start)
    
    res3.save("Report4.xlsx")
