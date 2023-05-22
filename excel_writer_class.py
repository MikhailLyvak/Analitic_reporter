from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Border, Side, NamedStyle, Font, numbers


class Report:
    # wb = Workbook()
    # ws = wb.active

    def __init__(self, data: list, report_name: str = None) -> None:
        self.data = data
        self.report_name = report_name
        self.repo_range = len(data) + 1
        self.wb = Workbook()
        self.ws = self.wb.active

    def write_data(
        self, row_start: int = 3, cell_start: int = 3, titles: list = None
    ) -> None:
        """Функція для внесення данних в Excel"""
        # Друкуємо заголовки
        for index, title in enumerate(titles):
            column = index + cell_start
            self.ws.cell(row=row_start, column=column, value=title)

        # Друкуємо дату
        for row_num, row in enumerate(self.data, start=row_start + 1):
            for col_num, value in enumerate(row, start=cell_start):
                self.ws.cell(row=row_num, column=col_num, value=value)

    def number_columns(self, row_start: int = 3, cell_start: int = 2) -> None:
        """Функція для нумерація рядків"""
        self.ws.column_dimensions["B"].width = 5
        self.ws.row_dimensions[row_start].height = 30
        for i in range(1, self.repo_range):
            # Робимо заголовок
            self.ws.cell(row=row_start, column=cell_start, value="№ п/п")
            # Робимо нумерацію
            self.ws.cell(row=i + 3, column=cell_start, value=i)

    def align_centre(self, start_row: int = 2, start_column: int = 1) -> None:
        """Відцентровує всі данні у файлі по центру"""
        for row in self.ws.iter_rows(min_row=start_row, min_col=start_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def borders(self, x_start: int = 3, y_start: int = 2) -> None:
        """Функція для обведення клітинок"""
        thick = Side(border_style="thin")

        for rows in self.ws.iter_rows(
            min_row=x_start, max_row=self.ws.max_row, min_col=y_start
        ):
            for cell in rows:
                cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)

    def zagolovok(self, place: list = ["B2", "E2"]):
        self.ws[place[0]].value = self.report_name
        self.ws.merge_cells(f"{place[0]}:{place[1]}")
        font = Font(
            name="Times New Roman", size=24, color="123456", italic=True, bold=True, condense=True
        )
        self.ws[place[0]].font = font
        self.ws.row_dimensions[2].height = 40

    def chart(
        self,
        data_row_start: int = 3,
        data_col_start: int = 4,
        data_row_end: int = 4,
        data_col_end: int = 5,
        pillar_names_col_start: int = 3,
        pillar_names_row_start: int = 4,
        pillar_names_row_end: int = 6,
        chart_placement: str = "D11"
    ) -> None:
        """Функція для створення графіків"""
        # TODO: ===========================
        # Приклад використання chart(1, 2, 3, 4, 5, 6, 7, 8
        # 1) - рядок де починаються данні(числа) ! з назвою колонок ВАЖЛИВО !
        # 2) - стовпець де почианються данні
        # 3) - рядок де закінчуються данні для графіка
        # 4) - стовпець де закінчуються данні для графіка
        # 5, 6, 7) - старт стовпець|старт рядок|канцевий рядок з назвами стовбчиків в графіку
        # 8) - Положення графіку у форматі "D11"

        # Тут кладемо інформацію (цифри і назви місяців)
        data = Reference(
            self.ws,
            min_col=data_col_start,
            min_row=data_row_start,
            max_col=data_col_end,
            max_row=data_row_end,
        )
        # Тут кладемо імені яким знизу підпишуться стовбчики
        categories = Reference(
            self.ws,
            min_col=pillar_names_col_start,
            min_row=pillar_names_row_start,
            max_row=pillar_names_row_end,
        )

        # Create the chart object
        chart = BarChart()
        # titles_from_data=True з data бере строки (тут це місяці і називає ними рядки)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = "Sales department"
        chart.x_axis.title = "Name"
        chart.y_axis.title = "Sales"
        chart.width = 33
        chart.height = 12

        # Тут ми просто ствоюємо графік і розміщаємо його де нам потрібно
        self.ws.add_chart(chart, chart_placement)
        
    def design_titles(self, row_number:int = 3) -> None:
        font = Font(bold=True)
        for cell in self.ws[row_number]:
            cell.font = font
    
    def money_format(
        self,
        start_row: int,
        start_col: int,
        end_row: int = None,
        end_col: int = None
    ) -> None:
        for row in self.ws.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            for cell in row:
                cell.number_format = '#,##0.00'
    
    def format_cols_width(
        self,
        column_width: int,
        start_col: str,
        end_col: str
    ) -> None:
        for column in self.ws[start_col:end_col]:
            self.ws.column_dimensions[column[0].column_letter].width = column_width
    
    def name_align_left(
        self,
        start_row: int,
        start_col: int,
        max_col: int = None
    ) -> None:       
        for row in self.ws.iter_rows(min_row=start_row, min_col=start_col, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
                
    def print_sum(
        self,
        data: list,
        col_start: int = 6
    ) -> None:
        border = Border(
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )

        for index, value in enumerate(data):
            col = col_start + index
            row = self.repo_range + 3
            cell = self.ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
        
        font = Font(bold=True)
        for cell in self.ws[self.repo_range + 3]:
            cell.font = font
