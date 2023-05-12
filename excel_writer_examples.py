from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Border, Side, NamedStyle, Font


wb = Workbook()


ws = wb.active

# ws.cell(row=2, column=2, value="№")
ws.cell(row=2, column=3, value='Name')
ws.cell(row=2, column=4, value='Novenber')
ws.cell(row=2, column=5, value='December')

values = [
    ['John', 25, 100],
    ['Alice', 30, 50],
    ['Bob', 20, 79],
]


def number_columns(
    start: int = 2,
    range_s: int = 1,
    range_e: int = 4
) -> None:
    """Функція для нумерація рядків"""
    for i in range(range_s, range_e):
        # Робимо заголовок
        ws.cell(row=2, column=2, value="№")
        # Робимо нумерацію
        ws.cell(row=i + 2, column=start, value=i)
    

def write_date() -> None:
    """Функція для внесення данних в Excel"""
    for row_num, row in enumerate(values, start=3):
        for col_num, value in enumerate(row, start=3):
            ws.cell(row=row_num, column=col_num, value=value)


def align_centre() -> None:
    """Відцентровує всі данні у файлі по центру"""   
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


def borders() -> None:
    thick = Side(border_style="thin")
    medium = Side(border_style="medium")
    
    
    for rows in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2):
        for cell in rows:
            cell.border = Border(outline=False, top=thick, left=thick, right=thick, bottom=thick, start=None, end=None)
    

def zagolovok():
    ws.column_dimensions['B'].width=10
    ws['B1'].value="Звіт з практики за "
    ws.merge_cells('B1:E1')
    font=Font(name='Times New Roman',size=24,color='123456',bold=True,condense=True)
    ws['B1'].font=font







def chart() -> None:
    """Функція для створення графіків"""
    # Тут кладемо інформацію (цифри і назви місяців)
    data = Reference(ws, min_col=4, min_row=2, max_col=5, max_row=5)
    # Тут кладемо імені яким знизу підпишуться стовбчики
    categories = Reference(ws, min_col=3, min_row=3, max_row=5)

    # Create the chart object
    chart = BarChart()
    # titles_from_data=True з data бере строки (тут це місяці і називає ними рядки)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = "Sales department"
    chart.x_axis.title = "Name"
    chart.y_axis.title = "Sales"

    # Тут ми просто ствоюємо графік і розміщаємо його де нам потрібно
    ws.add_chart(chart, "D11")

zagolovok()
number_columns()
write_date()
align_centre()
borders()
chart()
wb.save("test.xlsx")
